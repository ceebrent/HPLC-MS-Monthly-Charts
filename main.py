import itertools
import os
import shutil
import sys
import numpy as np
import pandas as pd
from openpyxl import load_workbook
import errno
import xlsxwriter
import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Style, Font, PatternFill
from tkinter import *
from tkinter import filedialog
import glob
import csv


def silent_remove(filename):
    try:
        os.remove(filename)
    except OSError as e:  # this would be "except OSError, e:" before Python 2.6
        if e.errno != errno.ENOENT:  # errno.ENOENT = no such file or directory
            raise  # re-raise exception if a different error occurred

def text_to_df(path_to_folder):
    batch_folder = os.path.dirname(path_to_folder)
    list_of_files = glob.glob(os.path.join(path_to_folder, '*.txt'))

    
    ##For each csv file in list of files we append them to the end
    df = pd.concat((pd.read_csv(f, delimiter='\t') for f in list_of_files))

    df.replace(['BENZOYLECGONINE-COCAINE METABOLITE 1', 'JWH-073-N-Butanoic acid/JWH-018-5OHpentyl 1'],
               ['BENZOYLECGONINE-COCAINE 1', 'JWH-073-N-Butanoic/JWH-018 1'], inplace=True)
    df['Component Name'] = df['Component Name'].str.replace('/', '-')
    df['Calculated Concentration'] = pd.to_numeric(df['Calculated Concentration'], errors='coerce').round(2)        
    columns_to_keep = ['Sample Name', 'Sample Type', 'Component Name', 'Actual Concentration',
                       'Calculated Concentration', 'Inst.']
    return df[columns_to_keep]

def group_sort(df, path_to_folder):
    df = df[df['Component Name'].str.endswith('1')]
    df.sort_values(['Component Name', 'Sample Name', 'Inst.'], inplace=True)
    df.reset_index(inplace=True, drop=True)
    length = df.shape[0]
    df_name_modify = df
    df['Calculated Concentration'].fillna(0,inplace=True)
            
    csv_file = os.path.join(path_to_folder, 'data.csv')
    new_csv_file = os.path.join(path_to_folder, 'data_m.csv')
    silent_remove(new_csv_file)
    silent_remove(csv_file)
    df.to_csv(csv_file)
    num_patients = []
    with open(csv_file, 'r') as infile, open(new_csv_file, 'a', newline='') as outfile:
        reader = csv.reader(infile)
        writer = csv.writer(outfile)
        header = next(reader)
        sample_type = header.index('Sample Type')
        sample_name = header.index('Sample Name')
        writer.writerow(header)
        for row in reader:
            if row[sample_type] == 'Unknown':
                try:
                    row[sample_name] = int(row[sample_name][-2:].strip(' '))
                    writer.writerow(row)
                    num_patients.append(row[sample_name])
                except ValueError:
                    writer.writerow(row)
            else:
                writer.writerow(row)
    unique_num_patients = set(num_patients)
    unique_inst = [str(x) for x in df['Inst.'].unique()]
    silent_remove(csv_file)
    return df['Component Name'].unique(), unique_inst, new_csv_file, unique_num_patients
def get_cutoffs(csv_file):

    lower_cutoff = {}
    upper_cutoff = {}
    drug_dict = {'lower':lower_cutoff,
                 'upper':upper_cutoff}
    with open(csv_file, 'r') as infile:
        reader = csv.reader(infile)
        header = next(reader)
        sample_name = header.index('Sample Name')
        component_name = header.index('Component Name')
        actual_conc = header.index('Actual Concentration')
        for row in reader:
            if row[sample_name].lower() == 'cal 2' and row[component_name] not in lower_cutoff:
                lower_cutoff[row[component_name]] = float(row[actual_conc])
            if row[sample_name].lower() == 'cal 5' and row[component_name] not in upper_cutoff:
                upper_cutoff[row[component_name]] = float(row[actual_conc])

    return drug_dict
            
def write_styles(csv_file, path_to_folder, drug_names,
                 instruments, num_of_patients):
    out_xlsx = os.path.join(path_to_folder, 'Comparison_{inst1}_to_{inst2}.xlsx'.format(inst1=instruments[0],
                                                                                        inst2=instruments[1]))
    silent_remove(out_xlsx)
    summary_csv = os.path.join(path_to_folder, 'Summary.csv')
    silent_remove(summary_csv)

    with open(summary_csv, 'w', newline='') as outfile:
        writer = csv.writer(outfile)
        writer.writerow(['Analyte', 'Sample', 'Reason'])
    
    wb = openpyxl.Workbook()
    for drugs in range(len(drug_names)):
        wb.create_sheet(index=drugs, title=drug_names[drugs])

    drug_sheets = wb.get_sheet_names()
    drug_sheets.remove('Sheet')
    drug_cutoffs = get_cutoffs(csv_file)

    """Borders formatting"""
    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
    medium_border = Border(left=Side(style='medium'), 
                     right=Side(style='medium'), 
                     top=Side(style='medium'), 
                     bottom=Side(style='medium'))

    """Fill Color Formatting"""
    redFill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    yellowFill = PatternFill(start_color='F3FF00', end_color='F3FF00', fill_type='solid')
    greenFill = PatternFill(start_color='78FF00', end_color='78FF00', fill_type='solid')
    whiteFill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    
    with open(csv_file, 'r') as infile:
        reader = csv.reader(infile)
        reader_header = next(reader)
        header = ['Sample Name', instruments[0], instruments[1], '% Diff.']        
        for sheet in drug_sheets:
            if sheet == 'Sheet':
                pass
            ws = wb.get_sheet_by_name(sheet)
##          Write and style title name cell 
            wb[sheet].append([sheet])
            ws.cell(row=1, column=1).border = medium_border
            ws.cell(row=1, column=1).font = Font(bold=True, size=18)

##          Write and style header  
            wb[sheet].append(header)
            for col in range(1,len(header)+1):
                ws.cell(row=2, column=col).border = thin_border
                ws.cell(row=2, column=col).font = Font(bold=True)

##           Begin to write Sample names and machine calc concs.

        
            for sample in range(1, len(num_of_patients)+1):
                ws.cell(row=sample+2, column=1).value = sample
                sample_name = reader_header.index('Sample Name')
                inst_data = reader_header.index('Inst.')
                calc_conc = reader_header.index('Calculated Concentration')
                comp_name = reader_header.index('Component Name')
                inst1_value = 0
                inst2_value = 0
                diff_cell = ws.cell(row = sample + 2, column=4)
                for row in reader:
                    try:
                        if (int(row[sample_name]) == sample and row[comp_name] == sheet):
                            if row[inst_data] == instruments[0]:
                                
                                inst1_value = float(row[calc_conc])

                                if inst1_value < drug_cutoffs['lower'][sheet]:
                                    inst1_value = 0
  
                                ws.cell(row = sample+2, column=2).value = inst1_value
                                
                            elif row[inst_data] == instruments[1]:
                                inst2_value = float(row[calc_conc])

                                if inst2_value < drug_cutoffs['lower'][sheet]:
                                    inst2_value = 0

                                ws.cell(row = sample+2, column=3).value = inst2_value
                                
                            else:
                                pass
                    except ValueError:
                        pass
                try:
                    diff = round((100 * abs(inst1_value - inst2_value)/inst2_value), 2)
                    diff_cell.value = diff
                    if diff > 25:
                        diff_cell.fill = redFill
                        with open(summary_csv, 'a', newline='') as summary:
                            writer = csv.writer(summary)
                            writer.writerow([sheet, sample, diff])
                except ZeroDivisionError:
                    pass

                ##Begin Styling
                for col in range(1,len(header)+1):
                    for row in range(1, len(num_of_patients) + 3):
                        ws.cell(row=row, column=col).border = thin_border
                        
                ##Begin Logic + Styling
                if (inst1_value == 0 and inst2_value == 0):
                    diff_cell.value = 'Match'
                        
                elif (inst1_value > drug_cutoffs['upper'][sheet]) and (inst2_value > drug_cutoffs['upper'][sheet]):
                    diff_cell.value = 'Match: > High Cutoff'
                    diff_cell.fill = whiteFill
                        
                elif (inst1_value == 0 and inst2_value > 0) or (inst2_value == 0 and inst1_value > 0):
                    diff_cell.value = 'Mismatch'
                    diff_cell.fill = yellowFill
                    with open(summary_csv, 'a', newline='') as summary:
                        writer = csv.writer(summary)
                        writer.writerow([sheet, sample, 'Mismatch'])
                    
                infile.seek(1)
    wb.save(out_xlsx)


def make_file_dialog():
    root = Tk()
    root.withdraw()
    root.file_name = filedialog.askdirectory(parent=root,
                                             title='Choose a Comparison folder')

    if not root.file_name:
        sys.exit(0)
    df = text_to_df(root.file_name)
    drug_names, instruments, csv_file,num_patients = group_sort(df, root.file_name)
    write_styles(csv_file, root.file_name, drug_names, instruments, num_patients)
    silent_remove(csv_file)

if __name__ == "__main__":
    make_file_dialog()

