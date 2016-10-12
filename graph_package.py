import calendar
import itertools
import os
import shutil
import sys
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from matplotlib.backends.backend_pdf import PdfPages
import openpyxl
from openpyxl import load_workbook
import errno
import xlsxwriter
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Style, Font


plt.style.use('ggplot')


def flip(items, ncol):
    return itertools.chain(*[items[i::ncol] for i in range(ncol)])

"""Make this take a lab name from PYQT since the dictionary is already there
    and add the Lab name to top of graph as well"""

def silent_remove(filename):
    try:
        os.remove(filename)
    except OSError as e:  # this would be "except OSError, e:" before Python 2.6
        if e.errno != errno.ENOENT:  # errno.ENOENT = no such file or directory
            raise  # re-raise exception if a different error occurred

def make_chart(data_csv):
    data = pd.read_csv(data_csv, encoding='utf-8-sig', low_memory=False)
    df = pd.DataFrame(data)
    df.sort_values(['Component Name', 'Sample Name'], inplace=True)
    out_csv_file = os.path.join(os.path.dirname(data_csv), 'Precision_Validation.xlsx')                          
    silent_remove(out_csv_file)
    xlsxwriter.Workbook(out_csv_file)
    list_of_df = []
    grouped = df.groupby(['Component Name'])
    df_counter = 0
    high_counter = 5
    low_counter = 0

    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
    medium_border = Border(left=Side(style='medium'), 
                     right=Side(style='medium'), 
                     top=Side(style='medium'), 
                     bottom=Side(style='medium'))
    
    for drug in grouped.size().index:
        pd.options.mode.chained_assignment = None
        drug_name = drug[:-2]
        if(drug_name == 'Modafinil'):
            pass
        dict_to_excel = {}
        drug_group = grouped.get_group(drug)
        number_of_days = list(set(drug_group['Day']))
##        drug_group['Calculated Concentration'] = drug_group['Calculated Concentration'].convert_objects(convert_numeric=True).round(2)
        drug_group['Calculated Concentration'] = pd.to_numeric(drug_group['Calculated Concentration'], errors='coerce').round(2)
        day_dfs = []
        
        for days in number_of_days:
            out_dict = {}
            day = drug_group[drug_group['Day'] == days]
            headers = drug_group.dtypes.index
            day_header = 'Day {day}'.format(day=days)
            out_dict[(drug_name, day_header, 'Calc. Conc. (ng/mL)')] = day['Calculated Concentration']##.convert_objects(convert_numeric=True).round(2)
            out_dict[(drug_name, day_header, 'Sample Name')] = day['Sample Name']
            df_to_excel = pd.DataFrame.from_dict(out_dict)
            df_to_excel.reset_index(inplace=True, drop=True)
            df_to_excel = df_to_excel[[(drug_name, day_header, 'Sample Name'), (drug_name, day_header, 'Calc. Conc. (ng/mL)')]]
            df_to_excel.set_index((drug_name, day_header, 'Sample Name'), inplace=True)
            df_to_excel.index.rename('Sample Name', inplace=True)
            day_dfs.append(df_to_excel)

        result = day_dfs[0]
        for x in range(1, len(day_dfs)):
            result = pd.concat([result, day_dfs[x]], axis=1, join_axes=[result.index])

        """Being mean and std calculations"""
        high_qc_mean = drug_group[drug_group['Sample Name'] == 'HIgh QC']['Calculated Concentration'].mean().round(2)
        low_qc_mean = drug_group[drug_group['Sample Name'] == 'Low QC']['Calculated Concentration'].mean().round(2)
        high_qc_std = drug_group[drug_group['Sample Name'] == 'HIgh QC']['Calculated Concentration'].std().round(2)
        low_qc_std = drug_group[drug_group['Sample Name'] == 'Low QC']['Calculated Concentration'].std().round(2)
        high_qc_rsd = (high_qc_std / high_qc_mean * 100).round(2)
        low_qc_rsd = (low_qc_std / low_qc_mean * 100).round(2)
        if(low_counter == 0):
            low_counter = int(5 + len(result)/2)
        
        mean_column = 'G'
        std_column = 'H'
        rsd_column = 'I'

        high_qc_mean_header_loc = "{column}{row}".format(column=mean_column, row=high_counter)
        high_qc_std_header_loc = "{column}{row}".format(column=std_column, row=high_counter)
        high_qc_rsd_header_loc = "{column}{row}".format(column=rsd_column, row=high_counter)

        low_qc_mean_header_loc = "{column}{row}".format(column=mean_column, row=low_counter)
        low_qc_std_header_loc = "{column}{row}".format(column=std_column, row=low_counter)
        low_qc_rsd_header_loc = "{column}{row}".format(column=rsd_column, row=low_counter)

        high_qc_mean_data_loc = "{column}{row}".format(column=mean_column, row=high_counter+1)
        high_qc_std_data_loc = "{column}{row}".format(column=std_column, row=high_counter+1)
        high_qc_rsd_data_loc = "{column}{row}".format(column=rsd_column, row=high_counter+1)

        low_qc_mean_data_loc = "{column}{row}".format(column=mean_column, row=low_counter+1)
        low_qc_std_data_loc = "{column}{row}".format(column=std_column, row=low_counter+1)
        low_qc_rsd_data_loc = "{column}{row}".format(column=rsd_column, row=low_counter+1)

        
        
        book = load_workbook(out_csv_file)
        writer = pd.ExcelWriter(out_csv_file, engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        result.to_excel(writer, startrow=df_counter, startcol=0)
        """Begin to write means and stds into workbook"""
    
        book.active[high_qc_mean_header_loc] = "High QC Mean"
        book.active[high_qc_std_header_loc] = "High QC Std."
        book.active[high_qc_rsd_header_loc] = "High QC %RSD"

        book.active[high_qc_mean_data_loc] = high_qc_mean
        book.active[high_qc_std_data_loc] = high_qc_std
        book.active[high_qc_rsd_data_loc] = high_qc_rsd

        book.active[low_qc_mean_header_loc] = "Low QC Mean"
        book.active[low_qc_std_header_loc] = "Low QC Std."
        book.active[low_qc_rsd_header_loc] = "Low QC %RSD"

        book.active[low_qc_mean_data_loc] = low_qc_mean
        book.active[low_qc_std_data_loc] = low_qc_std
        book.active[low_qc_rsd_data_loc] = low_qc_rsd

        for columns in range(1,7):
            for rows in range(10):
                book.active.cell(row=high_counter + rows, column=columns).border = thin_border

        for columns in range(7,10):
            book.active.cell(row=high_counter, column=columns).font = Font(bold=True)
            book.active.cell(row=low_counter, column=columns).font = Font(bold=True)
            
            for rows in range(2):
                book.active.cell(row=high_counter + rows, column=columns).border = thin_border
                book.active.cell(row=low_counter + rows, column=columns).border = thin_border
        
        
        writer.save()
        df_counter = df_counter + 16
        high_counter = len(result) + 6 + high_counter
        low_counter = len(result)+ 6 + low_counter
